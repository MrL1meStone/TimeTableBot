import os
import io
import openpyxl
import json
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import datetime
def get_schedule(day, course, group, redirection=False):
    now = datetime.datetime.now()
    # Смещение: 1 день и 5 часов
    offset = datetime.timedelta(days=(0 + (1 * day=="Tomorrow")), hours=5)
    # Время со смещением
    new_time = now + offset
    perm_time = str(new_time).split()[0].split('-')[2]+"."+str(new_time).split()[0].split('-')[1]+"."+str(new_time).split()[0].split('-')[0]
    perm_time = '24.09.2025'
    path, file1 = 'C:\\Users\\Валентин\\TimeTableBot\\downloaded_xlsx', ""
    for fff in os.listdir(path):
        if 'Расписание' in fff:
            f1, f2, f3, f4 = fff.split()[4].split('-')[0].split(".")[1], fff.split()[4].split('-')[1].split(".")[1], fff.split()[4].split('-')[0].split(".")[0], fff.split()[4].split('-')[1].split(".")[0]
            if perm_time.split(".")[1] == f1 or perm_time.split(".")[1] == f2:
              if f3 <= perm_time.split(".")[0] <= f4:
                if fff.split()[1] == course: # course = '1-2' либо '3-4'
                    file1 = fff
                    print("Файл найден")
              elif not(f3 <= perm_time.split(".")[0] <= f4) and redirection == True:
                  return "Извините, расписания нет. Скорее всего, вы хотите посмотреть расписание на внеурочный день (воскресенье и т.д.)"
    if file1 != "":
        print("идет парсинг")
        p = "".join((path+'\\'+file1).split("~$"))
        workbook = openpyxl.load_workbook(p)
        sheet, need_row = workbook.active, ''
        if day == "Today" or day == "Tomorrow":
            for row in range(7, 74):
                cell = sheet[f'A{row}']
                if str(perm_time) in str(cell.value):
                    need_row = cell.row
                    break
        for i in sheet.iter_cols():
          for j in i:
            if j.value == group:
                need_group = j
        all_days = {}
        if day == "Today" or day == "Tomorrow":
            schedule = []
            if need_row == 14:
                  schedule.append([sheet.cell(row=12, column=4).value, sheet.cell(row=12, column=6).value])
                  all_days[""] = schedule
                  schedule = []
            for row in sheet.iter_rows(min_row = 7):
                nums = []
                for cell in row:
                    if (need_group.column <= cell.column <= (need_group.column+4)) or (cell.column == 2) or (cell.column == 4):
                        if cell.value != None and cell.value != [] and (need_row+11) >= cell.row >= (need_row):
                            nums.append(cell.value + (')' * (cell.column == 2)))
                if nums != []:
                    schedule.append(nums)
            all_days[sheet.cell(row=need_row, column=1).value] = schedule
            return all_days
        elif day == "Weekly":
            week_days = [[12, 25], [26, 37], [38, 49], [50, 61], [62, 73]]
            keys = [14, 26, 38, 50, 62]
            for i in range(0, 5):
              schedule = []
              for row in sheet.iter_rows(min_row = week_days[i][0], max_row = week_days[i][1]):
                nums = []
                for cell in row:
                    if (need_group.column <= cell.column <= (need_group.column+4)) or (cell.column == 2) or (cell.column == 4):
                        if cell.value != None and cell.value != []:
                            nums.append(str(cell.value) + (')' * (cell.column == 2)))
                if nums != []:
                    schedule.append(nums)
              all_days[sheet.cell(row=int(keys[i]), column=1).value] = schedule
            return all_days
    else:
        print("файл не найден")
        return apload(day, course, group, redirection=True)
def apload(d, c, g):
    print("проводим аплоуд")
    folder_path = 'C:\\Users\\Валентин\\TimeTableBot\\downloaded_xlsx'
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path):
            os.remove(file_path)
    try:
        with open('schedule.json', 'r', encoding='utf-8') as file:
          data = json.load(file)
    except FileNotFoundError:
        data = []
        with open('schedule.json', 'w', encoding='utf-8') as file:
            json.dump(data, file)
    SERVICE_ACCOUNT_FILE = "C:\\Users\\Валентин\\Downloads\\coral-firefly-472118-d4-0e2e192450f3.json"
    FOLDER_ID = '1jfQFPUpOuv_tNLyQIoipPVsLQHv_H5UQ'
    SCOPES = ['https://www.googleapis.com/auth/drive.readonly']
    current_files = []
    credentials = service_account.Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=SCOPES)
    service = build('drive', 'v3', credentials=credentials)

    #  Запрос на получение списка файлов
    results = service.files().list(
    q=f"'{FOLDER_ID}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'", fields="files(id, name)").execute()
    #  Получение списка файлов
    files = results.get('files', [])
    if not files:
        return print("В папке нет XLSX-файлов.")
    # Создание папки
    os.makedirs('downloaded_xlsx', exist_ok=True)
    # Скачивание
    for file in files:
      file_id = file['id']
      file_name = file['name']
      document = {}
      current_files.append(file_name)
      if not(file_name in data):
        print("новый файл")
        request = service.files().get_media(fileId=file_id)
        file_path = os.path.join('downloaded_xlsx', file_name)
        print(f"Скачивание {file_name}...")
        with open(file_path, 'wb') as f:
            downloader = MediaIoBaseDownload(f, request)
            done = False
            while not done:
                status, done = downloader.next_chunk()
                print(f"Прогресс: {int(status.progress() * 100)}%")
        print(f"{file_name} скачан(а).")
    s = get_schedule(d, c, g)
    return s
p = get_schedule("Weekly", "1-2", 'АП2-924')
for i in p:
    print(str(i)+" ключ")
    for j in p[i]:
      print(*j)
