import json
import os
import openpyxl

from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
from datetime import datetime, timedelta,date


FOLDER_ID = os.environ.get("FOLDER_ID")
CREDENTIALS = os.environ.get("CREDENTIALS")

ROWS_FOR_WEEK_DAYS = ((12, 25), (26, 37), (38, 49), (50, 61), (62, 73))
HEADER_ROWS = (14, 26, 38, 50, 62)
BASE_PATH = 'downloaded_xlsx'
CREDS_PATH = 'bot_settings/credentials.json'

if not os.path.exists(CREDS_PATH):
	with open(CREDS_PATH,'w') as file:
		file.write(json.dump(CREDENTIALS))


# ==============================
# ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ
# ==============================

def _parse_date_range(date_range: str) -> tuple[date, date]:
	"""
	Принимает строку вида '22.09-27.09.2025' (dd.mm-dd.mm.yyyy)
	и возвращает кортеж двух дат (начало, конец).
	>>> _parse_date_range('22.09-27.09.2025')
	(datetime.date(2025, 9, 22), datetime.date(2025, 9, 27))
	"""
	start_date, end_date = date_range.split('-')
	end_date = datetime.strptime(end_date, '%d.%m.%Y').date()
	start_date = datetime.strptime(f'{start_date}.{end_date.year}', '%d.%m.%Y').date()
	return start_date, end_date


def _remove_all_files() -> None:
	"""Удаляет все файлы в указанной папке."""
	for filename in os.listdir(BASE_PATH):
		file_path = os.path.join(BASE_PATH, filename)
		if os.path.isfile(file_path):
			os.remove(file_path)


def _download_new_xlsx() -> None:
	"""Скачивает новые XLSX-файлы из Google Drive в target_folder."""
	scopes = ['https://www.googleapis.com/auth/drive.readonly']
	credentials = service_account.Credentials.from_service_account_file(CREDS_PATH, scopes=scopes)
	service = build('drive', 'v3', credentials=credentials)

	results = service.files().list(
		q=f"'{FOLDER_ID}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'",
		fields="files(id, name)"
	).execute()

	files = results.get('files', [])
	if not files:
		raise FileNotFoundError("В папке нет файлов")

	os.makedirs(BASE_PATH, exist_ok=True)

	for file in files:
		if (file['name'] not in os.listdir(BASE_PATH) and
				file['name'] != 'Расписание звонков.xlsx' and
				"График" not in file['name']):
			request = service.files().get_media(fileId=file['id'])
			out_path = os.path.join(BASE_PATH, file['name'])
			with open(out_path, 'wb') as f:
				downloader = MediaIoBaseDownload(f, request)
				done = False
				while not done:
					status, done = downloader.next_chunk()

# ==============================
# ОСНОВНАЯ ЛОГИКА
# ==============================

def get_schedule(modifier: str, course: int, group: str):
	"""
	Возвращает расписание:
	- day: 'Today', 'Tomorrow' или 'Weekly'
	- course: '1-2' или '3-4'
	- group: название группы
	"""
	if course not in range(1,4+1):
		raise ValueError('Некорректно выбран курс')

	if modifier not in ("Weekly","Today","Tomorrow"):
		raise ValueError("Неправильный модификатор поиска расписания")

	now = datetime.now().astimezone()
	offset = timedelta(days=(modifier == "Tomorrow"), hours=5) - now.utcoffset()
	current_time = now + offset
	current_time += timedelta(days= (modifier=='Weekly' and current_time.date().weekday()==6))
	current_date = current_time.date()
	if current_date.weekday()==6:
		raise ValueError("Расписания на воскресение не существует")

	filename = ""

	# Ищем файл с подходящими датами
	for file in os.listdir(BASE_PATH):
		_ , course_range, _ , _ , date = file.split(" ")
		print(course_range,date)
		if str(course) in course_range:
			date = date.replace("г..xlsx", "") # Удаление ненужных частей в дате
			start_date, end_date = _parse_date_range(date)
			if start_date <= current_date <= end_date:
				filename = file
				break
	print(filename,1)

	if not filename:
		return upload_and_retry(modifier, course, group)

	return parse_schedule_file(os.path.join(BASE_PATH, filename), modifier, group, current_date)


def parse_schedule_file(filepath: str, modifier: str, group: str, date: datetime.date)-> dict[str,list[dict[str,str]]]:
	"""
	Разбор XLSX с расписанием.
	"""
	workbook = openpyxl.load_workbook(filepath)
	sheet = workbook.active
	need_group = None
	for column in range(6,sheet.max_column+1,5): #
		if sheet.cell(row=10,column=column).value.replace(" /В/",'') == group:
			need_group = sheet.cell(row=10,column=column)
			break

	if not need_group:
		raise FileNotFoundError("Группа не найдена в файле")

	if modifier in ("Today", "Tomorrow"):
		return parse_single_day(sheet, need_group, date)
	else:
		return parse_weekly(sheet, need_group)


def parse_single_day(sheet, need_group, date: datetime.date) -> dict[str,list[dict[str,str]]]:
	"""Парсинг расписания на один день."""
	needed_row = None

	for row in range(14,74+1,12):
		if date.strftime("%d.%m.%Y") in str(sheet.cell(row=row, column=1).value):
			needed_row = row
			break
	if not needed_row:
		raise FileNotFoundError("Расписания на этот день не существует")

	day_schedule = []
	if date.weekday == 0:
		day_schedule.append({'time':'8.30-9.10',
		                     'subject':'Классный час',
		                     "cabinet": "Не написан",
		                     'teacher':'Не написан'})
	for i in range(0,12,2): #6 ячеек, по 2 строки в каждой
		if sheet.cell(needed_row + i, need_group.column).value is not None:
			day_schedule.append({
				'time': sheet.cell(needed_row + i, 4).value,
				'subject': sheet.cell(needed_row + i, need_group.column).value,
				'cabinet': sheet.cell(needed_row + i, need_group.column + 3).value,
				'teacher': sheet.cell(needed_row + i + 1, need_group.column).value})

	return {sheet.cell(row=needed_row, column=1).value:day_schedule}


def parse_weekly(sheet, need_group) -> dict[str,list[dict[str,str]]]:
	"""Парсинг расписания на всю неделю."""
	weekly = {}
	for row in range(14,74+1,12): # 6 дней по 6 ячеек, в каждой по 2 строки, даты с 14 строки
		date: str = sheet.cell(row=row,column=1).value[:10] # в 1 колонке написаны даты
		# и день недели в формате dd.mm.yyyy WEEKDAY
		day = parse_single_day(sheet,need_group,datetime.strptime(date,"%d.%m.%Y"))
		weekly.update(day)
	return weekly


def upload_and_retry(day: str, course: int, group: str):
	"""
	Удаляет старые файлы, скачивает новые и повторяет получение расписания.
	"""
	print("Проводим загрузку новых файлов...")
	_remove_all_files()
	_download_new_xlsx()
	return get_schedule(day, course, group)

def parse_schedule(schedule: dict) -> str:
	message=''
	for key,value in schedule.items():
		message+=f"🗓{key}<br>"
		for n,subject in enumerate(value):
			message+=(f"{n+1}){subject['subject']}<br>"
			          f"    Преподаватель: {subject['teacher']}<br>"
			          f"    Время: {subject['time']}<br>"
			          f"    Кабинет: {subject['cabinet']}<br>")
		message += '<br>'
	return message